#!/usr/bin/env python3
"""
Автоматизация за обществени поръчки - Tech Hardware
Полу-автоматичен workflow: EOP email -> анализ -> draft до Polycomp

Поддържа два режима:
  - DEMO (по подразбиране): използва примерни данни, не пипа реални услуги
  - LIVE: реален IMAP, реален Playwright скрейпър, реален OpenAI анализ
Режимът се управлява от config.yaml -> runtime.live_mode (true/false).
"""
import yaml
import os
import json
import re
from datetime import datetime
from pathlib import Path

# ── Пътища (работят и в PyInstaller bundle) ─────────────────────────────
from paths import resource_path, user_data_path  # local helper

CONFIG_PATH = user_data_path("config.yaml")
CONFIG_EXAMPLE = resource_path("config.example.yaml")
OUTPUT_DIR = user_data_path("output")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


def load_config():
    """Зарежда config.yaml ако съществува, иначе config.example.yaml."""
    path = CONFIG_PATH if CONFIG_PATH.exists() else CONFIG_EXAMPLE
    return yaml.safe_load(Path(path).read_text(encoding="utf-8"))


def _log(cb, msg):
    """Изпраща съобщение към GUI callback ако има, и принтира."""
    print(msg)
    if cb:
        try:
            cb(msg)
        except Exception:
            pass


# ── 1. Четене на имейли ─────────────────────────────────────────────────
def check_email(log=None):
    """Чете имейли от ЦАИС ЕОП и обработва намерените поръчки.

    Връща списък с proc_id-та, които са обработени.
    """
    config = load_config()
    live = config.get("runtime", {}).get("live_mode", False)
    processed = []

    if not live:
        _log(log, "→ DEMO режим: генерирам примерна поръчка вместо реален IMAP.")
        demo_url = "https://app.eop.bg/today/" + datetime.now().strftime("%H%M%S")
        pid = process_procurement(demo_url, log=log)
        if pid:
            processed.append(pid)
        return processed

    # LIVE режим
    try:
        from imap_tools import MailBox, AND
    except ImportError:
        _log(log, "✖ imap_tools не е инсталиран. pip install imap-tools")
        return processed

    eop = config["eop"]
    _log(log, f"→ Свързване към IMAP {eop['imap_server']} ...")
    with MailBox(eop["imap_server"]).login(eop["imap_user"], eop["imap_password"]) as mailbox:
        sender = eop.get("sender_filter", "eop.bg")
        for msg in mailbox.fetch(AND(from_=sender, seen=False), mark_seen=False):
            _log(log, f"Намерен имейл: {msg.subject} — {msg.date}")
            body = msg.html or msg.text or ""
            links = re.findall(r"https://app\.eop\.bg/today/\d+", body)
            for link in set(links):
                pid = process_procurement(link, log=log)
                if pid:
                    processed.append(pid)
    if not processed:
        _log(log, "Няма нови имейли/поръчки.")
    return processed


# ── 2-5. Обработка на една поръчка ──────────────────────────────────────
def process_procurement(url, log=None):
    config = load_config()
    live = config.get("runtime", {}).get("live_mode", False)
    proc_id = url.rstrip("/").split("/")[-1]
    out = OUTPUT_DIR / f"PROC-{proc_id}"
    out.mkdir(parents=True, exist_ok=True)

    _log(log, f"\n=== Обработвам {proc_id} ===")
    _log(log, f"URL: {url}")

    # 1. Сваляне на данни
    if live and config.get("runtime", {}).get("use_scraper", False):
        try:
            from scraper_eop import fetch_procurement_details
            _log(log, "→ Стартирам Playwright скрейпър...")
            scraped = fetch_procurement_details(url)
            procurement_data = {
                "id": proc_id,
                "url": url,
                "title": scraped.get("title", "Без заглавие"),
                "buyer": scraped.get("buyer", "—"),
                "budget_bgn": scraped.get("budget", 0),
                "deadline": scraped.get("deadline", ""),
                "cpv": scraped.get("cpv", ""),
                "documents": scraped.get("documents", []),
            }
        except Exception as e:
            _log(log, f"✖ Скрейпърът се провали ({e}); ползвам примерни данни.")
            procurement_data = _demo_data(proc_id, url)
    else:
        procurement_data = _demo_data(proc_id, url)

    # 2. Анализ на документи
    analysis = analyze_documents(procurement_data, config, log=log)

    # 3. Прилагане на филтрите
    decision = apply_filters(procurement_data, analysis, config)

    # 4. Запис
    (out / "data.json").write_text(json.dumps(procurement_data, ensure_ascii=False, indent=2), encoding="utf-8")
    (out / "analysis.json").write_text(json.dumps(analysis, ensure_ascii=False, indent=2), encoding="utf-8")
    (out / "decision.json").write_text(json.dumps(decision, ensure_ascii=False, indent=2), encoding="utf-8")

    if decision["proceed"]:
        generate_polycomp_email(procurement_data, analysis, out, config)
        generate_offer_template(procurement_data, analysis, out, config)
        _log(log, f"✅ ПОДХОДЯЩА — генерирани файлове в {out}")
    else:
        _log(log, f"❌ ПРОПУСНАТА: {decision['reason']}")

    return proc_id


def _demo_data(proc_id, url):
    return {
        "id": proc_id,
        "url": url,
        "title": "Доставка на лаптопи и интерактивни дисплеи за училище",
        "buyer": "ОУ Примерно",
        "budget_bgn": 45000,
        "deadline": "2026-06-25",
        "cpv": "30213100-6",
        "documents": [],
    }


# ── AI анализ ───────────────────────────────────────────────────────────
def analyze_documents(proc, config=None, log=None):
    config = config or load_config()
    live = config.get("runtime", {}).get("live_mode", False)
    use_ai = config.get("runtime", {}).get("use_ai", False)

    if live and use_ai and os.getenv("OPENAI_API_KEY"):
        try:
            from analyzer_ai import analyze_text
            _log(log, "→ AI анализ с OpenAI...")
            text = "\n".join(d.get("text", "") for d in proc.get("documents", [])) or proc.get("title", "")
            raw = analyze_text(text)
            data = json.loads(raw)
            data.setdefault("requires_iso20000", False)
            data.setdefault("requires_serial_numbers", False)
            data.setdefault("items", [])
            data["extracted_text"] = text[:2000]
            return data
        except Exception as e:
            _log(log, f"✖ AI анализът се провали ({e}); ползвам евристика.")

    # Евристичен/демо анализ
    sample_text = """
    Техническа спецификация:
    - Лаптоп 15.6", i5, 16GB RAM, 512GB SSD - 20 бр.
    - Интерактивен дисплей 75" 4K с тъч - 3 бр.
    Изисквания към участника: ISO 9001
    Офертата да съдържа техническо предложение без серийни номера.
    """
    # Уважаваме отрицание: "без серийни номера" => НЕ изисква
    requires_serial = bool(re.search(r"сериен номер|серийни номера", sample_text, re.I)) and \
        not re.search(r"без\s+серийни?\s+номера?", sample_text, re.I)
    return {
        "requires_iso20000": bool(re.search(r"ISO\s*20000|ISO/IEC\s*20000", sample_text, re.I)),
        "requires_iso9001": bool(re.search(r"ISO\s*9001", sample_text, re.I)),
        "requires_serial_numbers": requires_serial,
        "items": [
            {"name": 'Лаптоп 15.6" i5 16GB/512GB', "qty": 20, "unit": "бр."},
            {"name": 'Интерактивен дисплей 75" 4K тъч', "qty": 3, "unit": "бр."},
        ],
        "extracted_text": sample_text,
    }


# ── Филтри ──────────────────────────────────────────────────────────────
def apply_filters(proc, analysis, config=None):
    config = config or load_config()
    f = config["filters"]
    budget = proc.get("budget_bgn", 0)

    if analysis.get("requires_iso20000") and f.get("require_iso_20000_skip"):
        return {"proceed": False, "reason": "Изисква ISO 20000"}
    if analysis.get("requires_serial_numbers") and f.get("require_serial_numbers_skip"):
        return {"proceed": False, "reason": "Иска серийни номера в офертата"}
    if budget < f.get("min_budget_bgn", 0):
        return {"proceed": False, "reason": f"Бюджет {budget} лв < минимум {f['min_budget_bgn']}"}

    title = proc.get("title", "").lower()
    if not any(k in title for k in f.get("keywords_include", [])):
        return {"proceed": False, "reason": "Не е техника (липсва ключова дума)"}
    if any(k in title for k in f.get("keywords_exclude", [])):
        return {"proceed": False, "reason": "Съдържа изключваща дума"}

    return {"proceed": True, "reason": "Отговаря на всички филтри"}


# ── Генериране на имейл и оферта ────────────────────────────────────────
def generate_polycomp_email(proc, analysis, out_dir, config=None):
    config = config or load_config()
    items_text = "\n".join(
        f"- {i['name']} - {i['qty']} {i.get('unit', 'бр.')}" for i in analysis["items"]
    )
    email = f"""Здравейте Polycomp,

Моля за оферта за следната обществена поръчка:
Поръчка: {proc['title']}
Възложител: {proc['buyer']}
Краен срок: {proc['deadline']}
Бюджет: {proc['budget_bgn']} лв без ДДС
Линк: {proc['url']}

Техническа спецификация:
{items_text}

Моля за:
1. Най-добра цена за посочените количества
2. Наличност и срок на доставка
3. Гаранционни условия
4. Алтернативни предложения ако има

Благодаря!
{config['polycomp']['your_company_name']}
{config['polycomp']['your_contact']}
"""
    (out_dir / "email_to_polycomp.txt").write_text(email, encoding="utf-8")


def generate_offer_template(proc, analysis, out_dir, config=None):
    config = config or load_config()
    try:
        from openpyxl import Workbook
    except ImportError:
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Ценова оферта"

    ws["A1"] = "Обществена поръчка"; ws["B1"] = proc["title"]
    ws["A2"] = "Възложител"; ws["B2"] = proc["buyer"]
    ws["A3"] = "Бюджет"; ws["B3"] = proc["budget_bgn"]

    ws["A5"] = "Артикул"; ws["B5"] = "Количество"
    ws["C5"] = "Ед.цена Polycomp"; ws["D5"] = "Наша продажна"; ws["E5"] = "Общо"

    row = 6
    for item in analysis["items"]:
        ws[f"A{row}"] = item["name"]
        ws[f"B{row}"] = item["qty"]
        ws[f"E{row}"] = f"=C{row}*B{row}"
        row += 1

    ws[f"A{row + 1}"] = "Доставка"
    ws[f"D{row + 1}"] = f"{config['profitability']['shipping_cost_percent']}%"
    ws[f"A{row + 2}"] = "Марж цел"
    ws[f"D{row + 2}"] = f"{config['profitability']['min_margin_percent']}%"

    wb.save(out_dir / "oferta_template.xlsx")


# ── Рентабилност ────────────────────────────────────────────────────────
def calc_profit(budget, poly_price, shipping=0, overhead=0):
    """Чиста функция за калкулатора — връща dict с резултатите."""
    total_cost = poly_price + shipping + overhead
    margin_bgn = budget - total_cost
    margin_pct = (margin_bgn / budget * 100) if budget else 0
    return {
        "total_cost": total_cost,
        "margin_bgn": margin_bgn,
        "margin_pct": margin_pct,
    }


def check_profit(proc_id, log=None):
    out = OUTPUT_DIR / f"PROC-{proc_id}"
    data = json.loads((out / "data.json").read_text(encoding="utf-8"))
    _log(log, f"Проверка рентабилност за {proc_id}")
    _log(log, "→ Попълни цените в oferta_template.xlsx и стартирай отново")
    return data


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--check-email", action="store_true")
    parser.add_argument("--check-profit", type=str)
    args = parser.parse_args()

    if args.check_email:
        check_email()
    elif args.check_profit:
        check_profit(args.check_profit)
    else:
        print("Използвай: python automation.py --check-email")
